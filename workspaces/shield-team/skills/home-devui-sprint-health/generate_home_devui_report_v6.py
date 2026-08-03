#!/usr/bin/env python3
"""
Home DevUI Sprint Health Report Generator v6
Generates a 5-tab Excel report: Status Report, 8.25, 9.0, 9.1, No Fix Version
Uses fresh data from JSON files saved by the Kiro skill.
"""
import json
import re
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === Configuration ===
OUTPUT_FILE = "Home_DevUI_Sprint_Health_2026-07-24.xlsx"
JIRA_BASE = "https://disneyexperiences.atlassian.net/browse/"

# Milestones
MILESTONES = {
    "8.25": {"dev_complete": "Jul 2", "feature_complete": "Jul 10", "code_complete": "Jul 24", "all_closed": "Aug 6"},
    "9.0": {"dev_complete": "Jul 31", "feature_complete": "Aug 7", "code_complete": "Aug 21", "all_closed": "Sep 4"},
    "9.1": {"dev_complete": "Sep 4", "feature_complete": "Sep 14", "code_complete": "Sep 28", "all_closed": "Oct 12"},
}

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
STATUS_COLORS = {
    "Closed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "In Testing": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "Ready for testing": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "Final Review": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "Code Review": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "Peer Review": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "In Review": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "In Development": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "In Progress": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Ready for Build": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "In Triage": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "Open": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
}
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def extract_tag(summary):
    """Extract text between {} from summary."""
    match = re.search(r'\{([^}]+)\}', summary)
    return match.group(1) if match else ""


def determine_moved(ticket, version_tab):
    """Determine if ticket was moved from another version."""
    key = ticket.get("key", "")
    fv = ticket.get("fix_versions", "")
    # 9.0 tickets that were moved from 8.25 (Android bugs COREEXP-2016 through COREEXP-2152)
    if version_tab == "9.0":
        if "HKDL" in fv and "Story" in ticket.get("issue_type", ""):
            return "Moved from 9.1"
        # Android bugs originally on 8.25, moved to 9.0
        if key.startswith("COREEXP-"):
            try:
                num = int(key.split("-")[1])
                if 2016 <= num <= 2152:
                    return "Moved from 8.25"
            except ValueError:
                pass
        return ""
    return ""


def status_sort_key(status):
    """Sort tickets by status progression."""
    order = {
        "Closed": 0, "In Testing": 1, "Ready for testing": 1,
        "Final Review": 2, "Code Review": 3, "Peer Review": 3,
        "In Review": 3, "In Development": 4, "In Progress": 4,
        "Ready for Build": 5, "In Triage": 5, "Open": 6
    }
    return order.get(status, 7)


def write_ticket_tab(ws, tickets, version_tab, show_moved=True):
    """Write a ticket listing tab."""
    headers = ["Key", "Summary", "Status", "SP", "Priority", "Assignee", "Fix Versions", "Type", "Tag"]
    if show_moved:
        headers.append("Moved")

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Sort tickets
    tickets.sort(key=lambda t: status_sort_key(t.get("status", "Open")))

    # Write data
    for row_idx, t in enumerate(tickets, 2):
        key = t["key"]
        summary = t.get("summary", "")
        status = t.get("status", "")
        sp = t.get("sp", 0)
        priority = t.get("priority", "")
        assignee = t.get("assignee", "Unassigned")
        fix_versions = t.get("fix_versions", "")
        issue_type = t.get("issue_type", "")
        tag = extract_tag(summary)
        moved = determine_moved(t, version_tab) if show_moved else ""

        # Key with hyperlink
        cell = ws.cell(row=row_idx, column=1, value=key)
        cell.hyperlink = f"{JIRA_BASE}{key}"
        cell.font = Font(color="0563C1", underline="single")
        cell.border = THIN_BORDER

        # Summary
        cell = ws.cell(row=row_idx, column=2, value=summary)
        cell.border = THIN_BORDER

        # Status with color
        cell = ws.cell(row=row_idx, column=3, value=status)
        cell.border = THIN_BORDER
        if status in STATUS_COLORS:
            cell.fill = STATUS_COLORS[status]

        # SP
        cell = ws.cell(row=row_idx, column=4, value=sp if sp else 0)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

        # Priority
        cell = ws.cell(row=row_idx, column=5, value=priority)
        cell.border = THIN_BORDER

        # Assignee
        cell = ws.cell(row=row_idx, column=6, value=assignee)
        cell.border = THIN_BORDER

        # Fix Versions
        cell = ws.cell(row=row_idx, column=7, value=fix_versions)
        cell.border = THIN_BORDER

        # Type
        cell = ws.cell(row=row_idx, column=8, value=issue_type)
        cell.border = THIN_BORDER

        # Tag
        cell = ws.cell(row=row_idx, column=9, value=tag)
        cell.border = THIN_BORDER

        if show_moved:
            cell = ws.cell(row=row_idx, column=10, value=moved)
            cell.border = THIN_BORDER

    # Set column widths
    widths = [14, 70, 18, 6, 14, 22, 45, 10, 18]
    if show_moved:
        widths.append(18)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_status_report(ws, data_825, data_90, data_91, data_nofv):
    """Write the Status Report summary tab."""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    row = 1
    # Title
    cell = ws.cell(row=row, column=1, value="Home Dev UI — Sprint Health Report")
    cell.font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y')}")
    row += 1
    ws.cell(row=row, column=1, value="Source: Filter 112465 (Home_All_at_Program Level)")
    row += 2

    # Scope Summary
    cell = ws.cell(row=row, column=1, value="SCOPE SUMMARY")
    cell.font = Font(bold=True, size=12)
    row += 1

    scope_headers = ["Version", "Tickets", "Total SP", "Feature Complete", "All Closed"]
    for col, h in enumerate(scope_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    row += 1

    for version, data, ms in [
        ("8.25", data_825, MILESTONES["8.25"]),
        ("9.0", data_90, MILESTONES["9.0"]),
        ("9.1", data_91, MILESTONES["9.1"]),
        ("No FV", data_nofv, {"feature_complete": "—", "all_closed": "—"})
    ]:
        total_sp = sum(t.get("sp", 0) or 0 for t in data)
        ws.cell(row=row, column=1, value=version).border = THIN_BORDER
        ws.cell(row=row, column=2, value=len(data)).border = THIN_BORDER
        ws.cell(row=row, column=3, value=total_sp).border = THIN_BORDER
        ws.cell(row=row, column=4, value=ms.get("feature_complete", "—")).border = THIN_BORDER
        ws.cell(row=row, column=5, value=ms.get("all_closed", "—")).border = THIN_BORDER
        row += 1

    row += 2

    # Status Breakdown per version
    for version, data in [("8.25", data_825), ("9.0", data_90)]:
        cell = ws.cell(row=row, column=1, value=f"STATUS BREAKDOWN — {version}")
        cell.font = Font(bold=True, size=11)
        row += 1

        status_groups = {}
        for t in data:
            s = t.get("status", "Open")
            if s not in status_groups:
                status_groups[s] = {"count": 0, "sp": 0}
            status_groups[s]["count"] += 1
            status_groups[s]["sp"] += (t.get("sp", 0) or 0)

        sh = ["Status", "Tickets", "SP"]
        for col, h in enumerate(sh, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = SUBHEADER_FILL
            c.border = THIN_BORDER
        row += 1

        for status in sorted(status_groups.keys(), key=status_sort_key):
            g = status_groups[status]
            ws.cell(row=row, column=1, value=status).border = THIN_BORDER
            ws.cell(row=row, column=2, value=g["count"]).border = THIN_BORDER
            ws.cell(row=row, column=3, value=g["sp"]).border = THIN_BORDER
            row += 1

        total_sp = sum(t.get("sp", 0) or 0 for t in data)
        ws.cell(row=row, column=1, value="TOTAL").border = THIN_BORDER
        c = ws.cell(row=row, column=2, value=len(data))
        c.border = THIN_BORDER
        c.font = Font(bold=True)
        c = ws.cell(row=row, column=3, value=total_sp)
        c.border = THIN_BORDER
        c.font = Font(bold=True)
        row += 2

    # Assignee Load
    cell = ws.cell(row=row, column=1, value="ASSIGNEE LOAD (8.25 + 9.0)")
    cell.font = Font(bold=True, size=11)
    row += 1

    all_active = data_825 + data_90
    assignee_load = {}
    for t in all_active:
        a = t.get("assignee", "Unassigned")
        if a not in assignee_load:
            assignee_load[a] = {"tickets": 0, "sp": 0}
        assignee_load[a]["tickets"] += 1
        assignee_load[a]["sp"] += (t.get("sp", 0) or 0)

    ah = ["Developer", "Tickets", "SP"]
    for col, h in enumerate(ah, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = SUBHEADER_FILL
        c.border = THIN_BORDER
    row += 1

    for dev in sorted(assignee_load.keys(), key=lambda x: assignee_load[x]["sp"], reverse=True):
        load = assignee_load[dev]
        ws.cell(row=row, column=1, value=dev).border = THIN_BORDER
        ws.cell(row=row, column=2, value=load["tickets"]).border = THIN_BORDER
        ws.cell(row=row, column=3, value=load["sp"]).border = THIN_BORDER
        row += 1

    row += 2

    # Capacity & Sprint Projection
    cell = ws.cell(row=row, column=1, value="SPRINT PROJECTION")
    cell.font = Font(bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1, value="Team capacity: 3 devs × 8 SP = 24 SP/sprint")
    row += 1
    ws.cell(row=row, column=1, value="Sprint duration: 2 weeks (TxP Sprint naming)")
    row += 1

    remaining_825 = sum((t.get("sp", 0) or 0) for t in data_825 if t.get("status") not in ["Closed", "In Testing", "Ready for testing", "Final Review"])
    remaining_90 = sum((t.get("sp", 0) or 0) for t in data_90 if t.get("status") not in ["Closed", "In Testing", "Ready for testing", "Final Review"])

    ws.cell(row=row, column=1, value=f"8.25 remaining (pre-testing): {remaining_825} SP")
    row += 1
    ws.cell(row=row, column=1, value=f"9.0 remaining (pre-testing): {remaining_90} SP")
    row += 2

    # Target Dates (from Release Calendar updated Jul 23, 2026)
    cell = ws.cell(row=row, column=1, value="TARGET DATES (Release Calendar)")
    cell.font = Font(bold=True, size=11)
    row += 1

    milestone_headers = ["Milestone", "8.25", "9.0", "9.1", "Status 8.25", "Status 9.0"]
    for col, h in enumerate(milestone_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = SUBHEADER_FILL
        c.border = THIN_BORDER
    row += 1

    from datetime import date as dt_date
    today = dt_date(2026, 7, 24)
    milestones = [
        ("Exp Client Dev Complete", dt_date(2026, 7, 2), dt_date(2026, 8, 14), dt_date(2026, 9, 14)),
        ("Feature Complete", dt_date(2026, 7, 10), dt_date(2026, 8, 21), dt_date(2026, 9, 21)),
        ("Code Complete", dt_date(2026, 8, 11), dt_date(2026, 9, 4), dt_date(2026, 10, 5)),
        ("All Tickets Closed", dt_date(2026, 8, 21), dt_date(2026, 9, 21), dt_date(2026, 10, 19)),
    ]

    for name, d825, d90, d91 in milestones:
        days_825 = (d825 - today).days
        days_90 = (d90 - today).days
        status_825 = f"🔴 {abs(days_825)}d PAST" if days_825 < 0 else (f"🟡 {days_825}d" if days_825 <= 14 else f"🟢 {days_825}d")
        status_90 = f"🔴 {abs(days_90)}d PAST" if days_90 < 0 else (f"🟡 {days_90}d" if days_90 <= 14 else f"🟢 {days_90}d")
        ws.cell(row=row, column=1, value=name).border = THIN_BORDER
        ws.cell(row=row, column=2, value=d825.strftime("%b %d")).border = THIN_BORDER
        ws.cell(row=row, column=3, value=d90.strftime("%b %d")).border = THIN_BORDER
        ws.cell(row=row, column=4, value=d91.strftime("%b %d")).border = THIN_BORDER
        ws.cell(row=row, column=5, value=status_825).border = THIN_BORDER
        ws.cell(row=row, column=6, value=status_90).border = THIN_BORDER
        row += 1


def main():
    # Load data from JSON files
    with open("home_devui_825_data.json", "r") as f:
        data_825 = json.load(f)
    with open("home_devui_90_data.json", "r") as f:
        data_90 = json.load(f)
    with open("home_devui_91_data.json", "r") as f:
        data_91 = json.load(f)
    with open("home_devui_nofv_data.json", "r") as f:
        data_nofv = json.load(f)

    wb = Workbook()

    # Tab 1: Status Report
    ws_status = wb.active
    ws_status.title = "Status Report"
    write_status_report(ws_status, data_825, data_90, data_91, data_nofv)

    # Tab 2: 8.25 Tickets
    ws_825 = wb.create_sheet("8.25 Tickets")
    write_ticket_tab(ws_825, data_825, "8.25", show_moved=True)

    # Tab 3: 9.0 Tickets
    ws_90 = wb.create_sheet("9.0 Tickets")
    write_ticket_tab(ws_90, data_90, "9.0", show_moved=True)

    # Tab 4: 9.1 Tickets
    ws_91 = wb.create_sheet("9.1 Tickets")
    write_ticket_tab(ws_91, data_91, "9.1", show_moved=False)

    # Tab 5: No Fix Version
    ws_nofv = wb.create_sheet("No Fix Version")
    write_ticket_tab(ws_nofv, data_nofv, "nofv", show_moved=False)

    wb.save(OUTPUT_FILE)
    print(f"✅ Report saved: {OUTPUT_FILE}")
    print(f"   8.25: {len(data_825)} tickets, {sum(t.get('sp',0) or 0 for t in data_825)} SP")
    print(f"   9.0:  {len(data_90)} tickets, {sum(t.get('sp',0) or 0 for t in data_90)} SP")
    print(f"   9.1:  {len(data_91)} tickets, {sum(t.get('sp',0) or 0 for t in data_91)} SP")
    print(f"   No FV: {len(data_nofv)} tickets")


if __name__ == "__main__":
    main()
